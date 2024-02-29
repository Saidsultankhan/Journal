from rest_framework import views
from openpyxl import Workbook
from openpyxl.styles import Font, Side, Border, Alignment
from openpyxl.utils import get_column_letter
from django.http import HttpResponse
from src.apps.jounal.models import Pupil


class ExcelGenerateViewSet(views.APIView):

    def get(self, request):
        wb = Workbook()
        wb.remove(wb.active)
        
        pupils_sheet = wb.create_sheet('Ученики', 0)
        pupils_sheet.append(['№','Имя_ru', 'Имя_en', 'Имя_uz', 'Класс', 'Классный руководитель', 'Родители'])

        number = 1
            # for dairy_data in obj.dairy_pupil.values():
            #     print(dairy_data['mark'])
            # print(obj.dairy_pupil.values())
        for obj in Pupil.objects.prefetch_related('dairy_pupil', 'grade', 'parent').all():
            pupils_sheet.append(
                [
                    number, 
                    obj.name_ru, 
                    obj.name_en,
                    obj.name_uz, 
                    str(obj.grade),
                    obj.grade.teacher.user.username,
                     obj.parent.name_uz
                ]
            )
            number+=1

        max_lengths = {}
        first_iteration = True
        
        for row in pupils_sheet.iter_rows(
            min_row=1,
            max_row=pupils_sheet.max_row,
            min_col=1,
            max_col=pupils_sheet.max_column
            ):
            for cell in row:
                if first_iteration:
                    cell.style = 'Good'
                    cell.font = cell.font.copy(bold=True, size=14, color='000000')
                pupils_sheet[cell.coordinate].alignment = Alignment(horizontal='center')
                row[0].font = Font(bold=True, size=14)

                if cell.value:
                    length = len(str(cell.value))
                    if cell.column_letter not in max_lengths or length > max_lengths[cell.column_letter]:
                        max_lengths[cell.column_letter] = length
                cell.border = Border(
                    left=Side(border_style="thin", color='000000'),
                    right=Side(border_style="thin", color='000000'),
                    top=Side(border_style="thin", color='000000'),
                    bottom=Side(border_style="thin", color='000000')
                )
            first_iteration = False

        for col, length in max_lengths.items():
            pupils_sheet.column_dimensions[col].width = length + 10

        file_name = 'pupils_detail.xlsx'

        wb.save(file_name)

        with open(file_name, 'rb') as excel_file:
            response = HttpResponse(excel_file.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = f'attachment; filename={file_name}'
            return response
        